import pandas as pd
import telebot
import json
from telebot import types
import os
import openpyxl as ox
from openpyxl import load_workbook
import jinja2
import sqlite3
from data_msg import start_msg, get_instrument, supp, text_final, text_perenos
import shutil
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from dadata import Dadata
import sql_write



token = '6192798371:AAGaLj8lllkkx8k39hi85kgGpVQnUZKHZno'   # https://t.me/template_test2023_bot
token_secret = 'b56d81cab37bb66a3cf7771421ceef9bb56a7160'
token_ddata = '9a2b8ad77b01418de2ce8c8a47d27c13df23fd9d'
db_name = 'users.db'

nalog = ''
bot = telebot.TeleBot(token)
bot.set_my_commands([
    telebot.types.BotCommand("/start", "Перезапуск бота"),
    telebot.types.BotCommand("/help", "Обратная связь"),
])


def find_by_inn(inn):
    dadata = Dadata(token_ddata)
    result = dadata.find_by_id("party", inn)
    return result

def find_by_passp(number):
    dadata = Dadata(token_ddata, token_secret)
    result = dadata.clean("passport", number)
    return result

def send_mail(text, recip, message):
    passwd = 'UgHidpsKijegVHMi40mz'
    server = 'smtp.mail.ru'
    user = 'alfatron123rus@mail.ru'
    recipients = [recip]
    sender = 'alfatron123rus@mail.ru'
    subject =f'Сообщение от чат-бота маркетплейса'
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = 'Bot <' + sender + '>'
    msg['To'] = ', '.join(recipients)
    msg['Reply-To'] = sender
    msg['Return-Path'] = sender
    msg['X-Mailer'] = 'Python/'

    part_text = MIMEText(text, 'plain')
    maintype = 'application'
    subtype = 'octet-stream'
    file = MIMEBase(maintype, subtype)
    with open(f'files/{message.from_user.id}_res.xlsx', 'rb') as fp:
        file.set_payload(fp.read())
    encoders.encode_base64(file)
    file.add_header('Content-Disposition', 'attachment', filename=f'Расчет акции.xlsx')
    msg.attach(part_text)
    msg.attach(file)
    mail = smtplib.SMTP_SSL(server)
    mail.login(user, passwd)
    mail.sendmail(sender, recipients, msg.as_string())
    mail.quit()

def update_table(message, df_user):
    # ________добавление столбцов_________________

    A = 'Артикул товара'
    B = 'Бренд'
    C = 'Название'
    D = 'Категория'
    E = 'Артикул поставщика'
    F = 'Номенклатура (код 1С)'
    G = 'Последний баркод'
    H = 'Налог'
    I = 'Текущая скидка, в %'
    J = 'Текущая розничная цена'
    K = 'Скидка маркетплейса, %'
    L = 'Ваш %  комиссии маркетплейса'
    M = '% комисии, которую предложил маркетплейс при неучастии в акции'
    N = '% комисии, который предложил маркетплейс при участии в акции'
    O = 'cтоимость закупки и доставки 1 ед. товара до склада маркетплейса, в руб'
    P = 'стоимость доставки   товара до покупателя силами маркетплейса (по умолчанию 115р./ед)'
    Q = 'стоимость возврата от покупателя (вводит продавец или по умолчанию 33 р./ед)'
    R = 'стоимость хранения на маркетплейсе (вводит продавец или по умолчанию 10 р./ед. /мес)'
    S = 'Стоимость упаковки (вводит продавец или по умолчанию 10 р./ед)'
    T = 'цена товара на полке с предложенной WB скидкой'
    U = 'полная себестоимость товара'
    V = 'текущая цена на полке с вычетом комиссии WB, в рублях (до участия в акции, предложенной маркетплейсом)'
    W = 'размер  комиссии WB, в рублях (до участия в акции, предложенной маркетплейсом)'
    X = 'прибыль до налогообложения(до участия в акции предложенной маркетплейсом)'
    Y = 'чистая прибыль(до участия в акции предложенной маркетплейсом)'
    Z = 'размер  комиссии WB, в рублях (при отказе от участия в акции)'
    AA = 'прибыль до налогообложения(при отказе от  участия в акции)'
    AB = 'чистая прибыль(при отказе от участия в акции)'
    AC = 'размер  комиссии WB, в рублях (при  участии в акции)'
    AD = 'прибыль до налогообложения(при  участии в акции)'
    AE = 'чистая прибыль(при участии в акции)'


    ################################# Рассчет NM  #####################################
    df_user[T] = df_user[J]-df_user[J]*df_user[K]/100
    df_user[U] = df_user[O] + df_user[P] + df_user[Q] + df_user[R] + df_user[S]
    df_user[V] = df_user[J]-df_user[J]/100*df_user[L]-df_user[I]
    df_user[W] = df_user[J]/100*df_user[L]
    df_user[X] = df_user[V] - df_user[V] / 100 * df_user[I] - df_user[U]
    df_user[Y] = df_user[X] - df_user[X] * df_user[H] / 100
    df_user[Z] = df_user[J] / 100 * df_user[M]
    df_user[AA] = df_user[J] - df_user[Z] - df_user[U]
    df_user[AB] = df_user[AA] - df_user[AA] * df_user[H] / 100
    df_user[AC] = df_user[J] / 100 * df_user[N]
    df_user[AD] = df_user[T] - df_user[AC] - df_user[U]
    df_user[AE] = df_user[AD] - df_user[AD] * df_user[H] / 100


    writer = pd.ExcelWriter(f'files/{message.from_user.id}_res.xlsx', engine='xlsxwriter')
    df_user.to_excel(writer, 'Лист1', index=False)
    workbook = writer.book
    header_format2 = workbook.add_format({'valign': 'right'})
    sheet = writer.sheets['Лист1']
    sheet.set_column('A:AH', 25, header_format2)
    sheet.set_row(0, 30, cell_format=header_format2)
    writer.save()

    con = sqlite3.connect("users.db")
    cursor = con.cursor()
    perem = message.from_user.id

    cursor.execute('SELECT mail FROM users WHERE from_user_id = ?', [perem])
    mail_name = cursor.fetchone()
    con.commit()
    cursor.close()
    mail_name = str(mail_name[0])
    print(type(mail_name))

    send_mail(text=text_final, recip=mail_name, message=message)
    bot.send_document(message.chat.id, open(f'files/{message.from_user.id}_res.xlsx', 'rb'),
                      caption=text_perenos, reply_markup=main_menu())



def change_pr1(message, pr1, col):
    # ________Изменить % базовой комиссии _________________

    df_user = pd.read_excel(f'files/{message.from_user.id}_res.xlsx')
    df_user[col] = pr1
    update_table(message, df_user=df_user)


def update_spreadsheet(path: str, _df, starcol: int = 1, startrow: int = 1, sheet_name: str = "ToUpdate"):
    '''

    :param path: Путь до файла Excel
    :param _df: Датафрейм Pandas для записи
    :param starcol: Стартовая колонка в таблице листа Excel, куда буду писать данные
    :param startrow: Стартовая строка в таблице листа Excel, куда буду писать данные
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :return:
    '''

    wb = ox.load_workbook(path)
    for ir in range(0, len(_df)):
        for ic in range(0, len(_df.iloc[ir])):
            wb[sheet_name].cell(startrow + ir, starcol + ic).value = _df.iloc[ir][ic]
    wb.save(path)

def start_markup():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    itembtn1 = types.KeyboardButton('Просчитать акцию')
    itembtn2 = types.KeyboardButton('Проверить контрагента по ИНН')
    itembtn3 = types.KeyboardButton('Проверить паспорт руководителя')
    markup.add(itembtn1, itembtn2, itembtn3)
    return markup

def main_menu():
    markup = types.ReplyKeyboardMarkup(row_width=1)
    itembtn1 = types.KeyboardButton('Обновить данные в таблице')
    itembtn2 = types.KeyboardButton('Изменить Ваш % комиссии маркетплейса')
    itembtn3 = types.KeyboardButton('Изменить % комиcсии при неучастии в акции')
    itembtn4 = types.KeyboardButton('Изменить % комисcии при участии в акции')
    itembtn5 = types.KeyboardButton('Назад')
    markup.add(itembtn1, itembtn2, itembtn3, itembtn4, itembtn5)
    return markup

def main():
    list_status = ['new_user', 'get_file', 'buy_premium']
    @bot.message_handler(commands=['start'])
    def start_message(message):

        sql_write.add_users(message, list_status[0])
        bot.send_message(message.chat.id,
                         text=f"Здравствуйте, {message.from_user.first_name.format(message.from_user)}, {start_msg}️",
                         parse_mode='HTML', reply_markup=start_markup())

    @bot.message_handler(content_types=['text'])
    def commands(message):
        if message.text == "Просчитать акцию":
            msg = bot.send_message(message.chat.id, text="Введите адрес электронной почты, куда вам отправить результаты расчетов", reply_markup=telebot.types.ReplyKeyboardRemove())
            bot.register_next_step_handler(msg, get_mail)
        elif message.text == 'Проверить контрагента по ИНН':
            msg = bot.send_message(message.from_user.id, text=f"Введите ИНН контрагента для проверки в формате\n/find_inn 5404193928")
            bot.register_next_step_handler(msg, find_inn)
        elif message.text == 'Проверить паспорт руководителя':
            msg = bot.send_message(message.from_user.id, text=f"Введите серию и номер паспорта руководителя для проверки в формате\n/find_passp:0104 657474")
            bot.register_next_step_handler(msg, find_pasp)
        elif message.text == "Обновить данные в таблице":

            bot.send_document(message.chat.id, open(f'шаблон.xlsx', 'rb'),
                              caption=f"{message.from_user.first_name.format(message.from_user)}, {get_instrument}",
                             reply_markup=telebot.types.ReplyKeyboardRemove())
        elif message.text == "Изменить Ваш % комиссии маркетплейса":
            msg = bot.send_message(message.chat.id,
                             text=f"{message.from_user.first_name.format(message.from_user)}️, введите % вашей комиссии\n в формате /pr1 18",
                             parse_mode='HTML', reply_markup=start_markup())
            bot.register_next_step_handler(msg, input_proc)

        elif message.text == "Изменить % комиcсии при неучастии в акции":
            msg = bot.send_message(message.chat.id,
                             text=f"{message.from_user.first_name.format(message.from_user)}️, введите % комиcсии при неучастии в акции\n в формате /pr2 25",
                             parse_mode='HTML', reply_markup=start_markup())
            bot.register_next_step_handler(msg, input_proc)

        elif message.text == "Изменить % комисcии при участии в акции":
            msg = bot.send_message(message.chat.id,
                             text=f"{message.from_user.first_name.format(message.from_user)}️, введите % комисcии при участии в акции\n в формате /pr3 10",
                             parse_mode='HTML', reply_markup=start_markup())
            bot.register_next_step_handler(msg, input_proc)
        elif message.text == "Назад":
            msg = bot.send_message(message.chat.id,
                             text=f"Здравствуйте, {message.from_user.first_name.format(message.from_user)}, {start_msg}️",
                             parse_mode='HTML', reply_markup=start_markup())

        elif '/help' in message.text:
            bot.send_message(message.chat.id, text=supp)
    def find_inn(message):
        if '/find_inn' in message.text:
            res = find_by_inn(message.text.split(' ')[1])
            print(res)
            try:
                text = f"Название организации {res[0]['value']}\n" \
                    f"Статус {res[0]['data']['state']['status']}\n" \
                    f"ИНН {res[0]['data']['inn']}\n" \
                    f"ОГРН {res[0]['data']['ogrn']}\n"\
                    f"Адрес {res[0]['data']['address']['value']}" \
                    f""

                    # f""
                bot.send_message(message.from_user.id, text=text, reply_markup=start_markup())
                print(res)

            except Exception as ex:
                bot.send_message(message.from_user.id, text=f"Вы ввели неверный ИНН", reply_markup=start_markup())
            

    def find_pasp(message):
        if '/find_passp' in message.text:
            res = find_by_passp(message.text.split(':')[1])
            print(res)
            text = ''
            if res['qc'] == 0:
                print('Действующий паспорт')
                text = f"Все в порядке, паспорт действующий"
            elif res['qc'] == 10:
                text = 'Внимание! Недействительный паспорт'
            elif res['qc'] == 1:
                text = 'Вы ввели неправильный формат серии или номера'
            bot.send_message(message.from_user.id, text=text, reply_markup=start_markup())


    def input_proc(message):
        if '/pr1' in message.text:
            change_pr1(message=message, pr1=(float(message.text.strip().split(' ')[1])), col="Ваш %  комиссии WB")
        elif '/pr2' in message.text:
            change_pr1(message=message, pr1=(float(message.text.strip().split(' ')[1])), col="% комисии, которую предложил WB при неучастии в акции")
        elif '/pr3' in message.text:
            change_pr1(message=message, pr1=(float(message.text.strip().split(' ')[1])), col="% комисии, который предложил WB при участии в акции")

    def get_mail(message):
        if "@" in message.text:
            mail_adress = message.text.strip()
            con = sqlite3.connect("users.db")
            cursor = con.cursor()
            zapros = f"""
                       UPDATE users 
                       SET mail = '{mail_adress}'
                       WHERE from_user_id = '{message.from_user.id}'
                                  """

            cursor.execute(zapros)
            con.commit()
            cursor.close()
            bot.send_message(message.chat.id, text="Ваша почта успешно сохранена")
            bot.send_document(message.chat.id, open(f'шаблон.xlsx', 'rb'),
                              caption=f"{message.from_user.first_name.format(message.from_user)}, {get_instrument}",
                             reply_markup=telebot.types.ReplyKeyboardRemove())
        else:
            bot.send_message(message.chat.id, text="Вы неверно указали адрес почты")
            msg = bot.send_message(message.chat.id,
                                   text="Введите адрес электронной почты, куда вам отправить результаты расчетов")
            bot.register_next_step_handler(msg, get_mail)


    @bot.message_handler(content_types=['document'])
    def input_file(message):
        if 'xls' in message.document.file_name:
            bot.send_message(message.chat.id, text='Загружаем файл.... ждите')

            if not os.path.exists(f"files"):
                os.mkdir(f"files")

            file_info = bot.get_file(message.document.file_id)
            src = f"files/{message.from_user.id}.xlsx"
            downloaded_file = bot.download_file(file_info.file_path)

            with open(src, 'wb') as new_file:
                new_file.write(downloaded_file)
            df_user = pd.read_excel(f'files/{message.from_user.id}.xlsx')
            update_table(message, df_user=df_user)

        else:
            bot.send_message(message.chat.id, text='Вы загрузили файл неверного формата, загрузите пожалуйста заполненный шаблон в формате excel')

    bot.polling(none_stop=True)

if __name__ == '__main__':
    main()